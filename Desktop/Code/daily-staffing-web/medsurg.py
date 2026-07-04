"""
medsurg.py — Med/Surg turn-tracker logic (ms-turn skill), factored into
importable functions for the web app.

Part 1: append a day's assignment-grid export into the running
MedSurg_Assignments_Summary.xlsx (ms-hero logic).

Part 2: cross-reference today's roster against the summary to rank who is
most overdue for a med/surg assignment (ms-turn logic).
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Part 1 — extract med/surg assignments from a day's assignment-grid export
# ---------------------------------------------------------------------------

HEART = "♥"
ROOM_RE = re.compile(r"\d+[A-Za-z]?[♥]?")
SKIP_KW = ("FOLEY", "LINE", "NAME", "EXT", "SHIFT", "CENSUS", "SITTER", "DA|", "ER|", "ED|", "ICU|", "FC:")


def looks_like_rooms(cell):
    if not cell:
        return False
    s = str(cell)
    if any(k in s.upper() for k in SKIP_KW):
        return False
    t = ROOM_RE.findall(s)
    return len(t) >= 2 and not any(len(x) > 4 for x in t)


def find_rooms_row(ws):
    best, score = None, 0
    for r in range(6, 11):
        s = sum(1 for c in ws[r] if looks_like_rooms(c.value))
        if s > score:
            score, best = s, r
    return best


def extract_assignment_date(ws, fallback):
    for cell in ws[3]:
        v = cell.value
        if not v:
            continue
        m = re.search(r"(\d{1,2}/\s*\d{1,2}\s*/\s*\d{2,4})", str(v))
        if m:
            return re.sub(r"\s+", "", m.group(1))
    return fallback


def analyze_assignment(path):
    """Returns (date_label, [(name, [room_tokens]), ...]) for nurses whose
    entire room list has no heart marker (i.e. med/surg, not cardiac/tele)."""
    wb = load_workbook(path)
    ws = wb.active
    names = [c.value for c in ws[5]]
    rr = find_rooms_row(ws)
    if rr is None:
        return extract_assignment_date(ws, Path(path).name), []
    rooms = [c.value for c in ws[rr]]
    date_lbl = extract_assignment_date(ws, Path(path).name)
    medsurg = []
    for name, room_cell in zip(names, rooms):
        if not name or not room_cell:
            continue
        nm = re.sub(r"\s+", " ", str(name).strip().replace("\n", " / "))
        rm = str(room_cell).strip()
        if nm.upper() in ("RN", "") or not rm:
            continue
        if "SITTER" in nm.upper() or "FOLEY" in rm.upper():
            continue
        tokens = ROOM_RE.findall(rm)
        if not tokens or any(len(t) > 4 for t in tokens):
            continue
        if not any(HEART in t for t in tokens):
            medsurg.append((nm, tokens))
    return date_lbl, medsurg


# ---------------------------------------------------------------------------
# Persistent summary workbook
# ---------------------------------------------------------------------------

BY_DAY_HEADERS = ["Date", "Nurse", "Rooms", "# Rooms"]
NURSE_TALLY_HEADERS = ["Nurse", "Med/Surg Days", "Dates"]
DAILY_SUMMARY_HEADERS = ["Date", "# of Med/Surg Nurses", "Nurses"]


def ensure_summary_exists(path):
    path = Path(path)
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "By Day"
    ws1.append(BY_DAY_HEADERS)
    ws2 = wb.create_sheet("Nurse Tally")
    ws2.append(NURSE_TALLY_HEADERS)
    ws3 = wb.create_sheet("Daily Summary")
    ws3.append(DAILY_SUMMARY_HEADERS)
    for ws in (ws1, ws2, ws3):
        for cell in ws[1]:
            cell.font = Font(bold=True)
    wb.save(path)


def append_to_summary(new_records, summary_path):
    """new_records: [(date_lbl, nurse, rooms_csv, count), ...]"""
    ensure_summary_exists(summary_path)
    wb = load_workbook(summary_path)
    ws1 = wb["By Day"]
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name="Arial", size=11)
    lwrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    light = PatternFill("solid", start_color="F2F6FC")

    ws1.append([])
    toggle = True
    for date_lbl, nurse, rooms_csv, cnt in new_records:
        ws1.append([date_lbl, nurse, rooms_csv, cnt])
        row = ws1.max_row
        for cell in ws1[row]:
            cell.font = body_font
            cell.border = border
            cell.alignment = center if isinstance(cell.value, int) else lwrap
        if toggle:
            for cell in ws1[row]:
                cell.fill = light

    all_records = [(str(r[0]), str(r[1])) for r in ws1.iter_rows(min_row=2, values_only=True) if r[0] and r[1]]
    tally = Counter()
    days_by = defaultdict(list)
    for d, n in all_records:
        tally[n] += 1
        days_by[n].append(d)

    ws2 = wb["Nurse Tally"]
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 4):
            ws2.cell(r, c).value = None
    for i, (nurse, count) in enumerate(sorted(tally.items(), key=lambda x: (-x[1], x[0])), 2):
        ws2.cell(i, 1).value = nurse
        ws2.cell(i, 2).value = count
        ws2.cell(i, 3).value = ", ".join(days_by[nurse])

    ws3 = wb["Daily Summary"]
    by_date = defaultdict(list)
    for d, n in all_records:
        by_date[d].append(n)
    existing = set(str(ws3.cell(r, 1).value) for r in range(2, ws3.max_row + 1) if ws3.cell(r, 1).value)
    for date_lbl, nurse, _, _ in new_records:
        if date_lbl not in existing:
            nurses = by_date.get(date_lbl, [])
            ws3.append([date_lbl, len(nurses), ", ".join(nurses)])
            existing.add(date_lbl)

    wb.save(summary_path)


def append_assignment_file(assignment_path, summary_path):
    """Analyze one day's assignment export and append to the summary.
    Returns (date_label, new_records)."""
    date_lbl, medsurg = analyze_assignment(assignment_path)
    new_records = [(date_lbl, nm, ", ".join(tokens), len(tokens)) for nm, tokens in medsurg]
    if new_records:
        append_to_summary(new_records, summary_path)
    return date_lbl, new_records


def validate_summary_workbook(path):
    """Raises ValueError if `path` doesn't look like a MedSurg summary file."""
    wb = load_workbook(path, read_only=True)
    required = {"By Day", "Nurse Tally", "Daily Summary"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Not a MedSurg summary file — missing sheet(s): {', '.join(sorted(missing))}")


# ---------------------------------------------------------------------------
# Part 2 — turn analysis: cross-reference roster against the summary
# ---------------------------------------------------------------------------

ELIGIBLE = {"RN", "RES/BRK"}

NICK = {
    "AGUILAR, NORA NATALIE": "NORA", "ADHIKARI, BIMALA": "BIMALA",
    "BALGOS, DENNIS": "DENNIS", "BELY, JULIA": "JULIA",
    "BERANGO, KATHRYN": "KATHRYN", "BROTHERS, JOHN": "JOHN",
    "BURLAKOTI, MUNA": "MUNA", "CARPENTER, SUSANNA": "SUSANNA",
    "CHEN, AIHUA": "AIHUA", "CICHELLA, COURTNIE": "COURTNIE",
    "FORST, RYAN": "RYAN", "GAMBETTI III, ALBERT": "AL",
    "GAMBETTI, ALBERT": "AL", "GA-AS, REINETTE": "REINETTE",
    "GAAS, REINETTE": "REINETTE", "HABIBI, MARYAM": "MARYAM",
    "HECHONA, MEREMEL": "MEREMEL", "HENDERSON, NATALIE": "NATALIE",
    "HERNANDEZ, DONNA": "DONNA", "IMMEKER, REBECCA": "REBECCA",
    "JARDIANIANO, CATHERINE": "CATHERINE", "JARDINIANO, CATHERINE": "CATHERINE",
    "JOHN, JOMINA": "JOMINA", "JONES, STEVEN": "STEVEN",
    "KAUR, KAMALDEEP": "KAMALDEEP", "LOPEZ, DENISE": "DENISE",
    "MADUIKE, SAMANTHA": "SAMANTHA", "MAGNO, ROWENA": "ROWENA",
    "MCKEETH, CURTIS": "CURTIS", "OYANGOREN, MA SOCORRO": "MA",
    "ROBSON, KELLI": "KELLI", "RYAN, MALGORZATA": "GOSIA",
    "SHOQUIST, MICHAEL": "MICHAEL", "SIGL, JOCELYN": "JOCELYN",
    "TRAN, TRANG": "TRANG", "VARELA, CARMEN": "CARMEN",
    "VARGHESE, ANU": "ANU", "VARGHESE, ASHA": "ASHA",
    "VLASIK, MARIYA": "MARIYA", "WALTON, RAVEN": "RAVEN",
    "BOREHAM, JANA": "JANA",
}
STRIP = re.compile(r"\b(III|lll|II|ll|IV|JR|SR)\b", re.I)


def normalize_display_name(s):
    s = re.sub(r"\(.*?\)", "", str(s).strip().upper())
    s = re.sub(r"\s*,\s*", ", ", s)
    s = STRIP.sub("", s)
    return re.sub(r"\s+", " ", s).strip().rstrip(",").strip()


def nickname(name):
    nm = normalize_display_name(name)
    if nm in NICK:
        return NICK[nm]
    if "," in nm:
        last, rest = nm.split(",", 1)
        fw = rest.strip().split()[0] if rest.strip() else ""
        short = f"{last.strip()}, {fw}"
        if short in NICK:
            return NICK[short]
        return fw
    return nm.split()[0]


def parse_dt(s):
    m = re.search(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?", str(s))
    if m:
        mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3) or 2026)
        if yr < 100:
            yr += 2000
        try:
            return date(yr, mo, dy)
        except ValueError:
            return None
    return None


def roster_shift_date(ws):
    for r in range(8, 14):
        v = ws.cell(r, 1).value
        if v and "date" in str(v).lower():
            m = re.search(r"(\w+ \d{1,2},?\s*\d{4})|(\d{1,2}/\d{1,2}/\d{2,4})", str(v), re.I)
            if m:
                raw = m.group(0).replace(",", "").strip()
                for fmt in ("%B %d %Y", "%b %d %Y", "%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.strptime(raw, fmt).date()
                    except ValueError:
                        pass
    return date.today()


def read_turn_eligible_nurses(ws):
    """Scan the roster sheet for RN / RES-BRK employees (RN CHG excluded)."""
    data_start = next(
        (r for r in range(13, 50)
         if ws.cell(r, 3).value and str(ws.cell(r, 3).value).strip()
         and str(ws.cell(r, 3).value).strip() != "Employee Name"),
        15,
    )
    nurses = []
    for r in range(data_start, data_start + 40):
        name = ws.cell(r, 3).value
        profile = ws.cell(r, 6).value
        code = ws.cell(r, 10).value
        if not name or not str(name).strip():
            break
        nm, pr, cd = str(name).strip(), str(profile or "").strip(), str(code or "").upper()
        if pr == "RN CHG":
            continue
        is_p = "PRCPT" in cd or "PRCPT" in pr.upper()
        base = pr.replace(" PRCPT", "").strip()
        if base in ELIGIBLE or pr in ELIGIBLE:
            nurses.append((nm, pr, is_p))
    return nurses


def compute_turn(roster_path, summary_path):
    ensure_summary_exists(summary_path)
    wb = load_workbook(roster_path, data_only=True)
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active
    shift_date = roster_shift_date(ws)
    nurses = read_turn_eligible_nurses(ws)

    wb2 = load_workbook(summary_path, data_only=True)
    ws2 = wb2["By Day"]
    last_ms = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        key = str(row[1]).split("/")[0].strip()
        dt = parse_dt(str(row[0]))
        if dt and (key not in last_ms or dt > last_ms[key]):
            last_ms[key] = dt

    results = []
    for name, profile, is_p in nurses:
        nick = nickname(name)
        last_dt = last_ms.get(nick)
        if last_dt:
            da = (shift_date - last_dt).days
            results.append((name, profile, is_p, f"{last_dt.month}/{last_dt.day}", da))
        else:
            results.append((name, profile, is_p, "NEVER", 9999))
    results.sort(key=lambda x: (-x[4], x[0]))

    never_e = [(n, p) for n, p, pt, ld, _ in results if ld == "NEVER" and not pt]
    never_p = [(n, p) for n, p, pt, ld, _ in results if ld == "NEVER" and pt]
    had_it = [(n, p, ld, da) for n, p, pt, ld, da in results if ld != "NEVER"]

    return shift_date, results, never_e, never_p, had_it


def render_turn_sheet(wb, shift_date, results, never_e, never_p, had_it):
    if "MS Turn" in wb.sheetnames:
        del wb["MS Turn"]
    ws = wb.create_sheet("MS Turn")
    bold = Font(bold=True)

    ws.cell(row=1, column=1,
            value=f"{shift_date.month}/{shift_date.day}/{shift_date.year} Day Shift — Med/Surg Turn Tracker").font = bold

    headers = ["NURSE", "PROFILE", "LAST MED/SURG", "DAYS AGO"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h).font = bold

    r = 4
    for name, profile, is_p, ld, da in results:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=profile)
        ws.cell(row=r, column=3, value=ld)
        days_val = da if da < 9999 else "—"
        if is_p and isinstance(days_val, int):
            days_val = f"{days_val} [PRCPT]"
        elif is_p:
            days_val = f"{days_val} [PRCPT]"
        ws.cell(row=r, column=4, value=days_val)
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="── RECOMMENDATION ──────────────────────────────────────────────────").font = bold
    r += 1

    if never_e:
        ws.cell(row=r, column=1, value="Has NEVER had med/surg:")
        r += 1
        for n, p in never_e:
            ws.cell(row=r, column=1, value=f"  → {n} ({p})")
            r += 1

    if never_p:
        ws.cell(row=r, column=1, value="Also NEVER (PRCPT — charge nurse discretion):")
        r += 1
        for n, p in never_p:
            ws.cell(row=r, column=1, value=f"  → {n} ({p})")
            r += 1

    r += 1
    label = "Next in line:" if never_e else "All have had med/surg. Most overdue:"
    ws.cell(row=r, column=1, value=label)
    r += 1
    for n, p, ld, da in had_it[:4]:
        ws.cell(row=r, column=1, value=f"  → {n} ({p})  last: {ld}  ({da} days ago)")
        r += 1

    ws.column_dimensions["A"].width = 60
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 16
    return ws
