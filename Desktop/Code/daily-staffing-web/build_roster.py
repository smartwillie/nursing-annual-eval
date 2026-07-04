#!/usr/bin/env python3
"""
build_roster.py — Build an enriched daily nursing roster.

Combines a "Daily Roster by Shift" export with the unit's RN/CNA rotation
tracker. Output:
  - sequential numbering in column B (RN-class and CNA-class separate streams)
  - last HC + last float dates (RN-class) or last sit + last float (CNA-class)
    in columns M and O of the roster's existing layout
  - a "sit / float" sub-header divider row between the two sections
  - two helper sheets: 'RN HC List' and 'RN Float List'
"""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Profile / code classification
# ---------------------------------------------------------------------------

RN_CLASS = {"RN", "RES/BRK", "RN CHG"}
RN_NUMBER_NO_DATES = set()
CNA_CLASS = {"CNA"}
CNA_NUMBER_NO_DATES = {"SA(P)"}
NO_TRACK = {"US", "NSM", "MON TECH", "MGR"}

NO_NUMBER_CODES = {"LT DUTY"}
NO_DATES_CODES = {"PRCPT"}

# Per-employee date-lookup overrides. Keys are normalized (LAST, FIRST) tuples.
EMPLOYEE_OVERRIDES = {
    ("BHANGU", "PAVANDEEP"): {"skip_left"},   # only float, no sit
}

# Name aliases. Applied on BOTH the roster side and the tracker side so any
# spelling of one person collapses into the same canonical key.
NAME_ALIASES = {
    ("EIERMAN", "DESTINY"):           ("EIRMAN", "DESTINY"),
    ("HUELE", "EVAJOANNE"):           ("HUELE", "EVA"),
    ("MANALO", "ANNKRISTIE"):         ("MANALO", "ANN"),
    ("WILHELM", "NICHOLAS"):          ("WILHELM", "NICK"),
    ("WILHEM",  "NICHOLAS"):          ("WILHELM", "NICK"),
    ("WILHEM",  "NICK"):              ("WILHELM", "NICK"),
    ("MENDEZRIVERA", "VICTOR"):       ("MENDEZRIVERA", "ANDREW"),
    ("AGUILAR", "NORANATALIE"):       ("AGUILAR", "NORA"),
    ("CALVAN", "SUERTE"):             ("CALVAN", "SUE"),
}

_PAREN_RE = re.compile(r"\([^)]*\)")
_GEN_SUFFIX_KEYWORDS = {"JR", "SR"}


# ---------------------------------------------------------------------------
# Name normalization
# ---------------------------------------------------------------------------

def _is_generational_roman_suffix(s):
    if not s:
        return False
    s = s.upper()
    if s in {"II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"}:
        return True
    return 2 <= len(s) <= 4 and all(c in "IL" for c in s)


def normalize_name(raw):
    """Normalize 'BALGOS, DENNIS (PD)' → ('BALGOS', 'DENNIS')."""
    if not isinstance(raw, str):
        return None
    cleaned = _PAREN_RE.sub("", raw).strip()
    if not cleaned:
        return None
    if "," in cleaned:
        last, first = cleaned.split(",", 1)
    else:
        parts = cleaned.split()
        if len(parts) < 2:
            return None
        last, first = parts[0], " ".join(parts[1:])
    last_tokens = last.strip().split()
    if len(last_tokens) > 1:
        tail = last_tokens[-1].upper()
        if tail in _GEN_SUFFIX_KEYWORDS or _is_generational_roman_suffix(tail):
            last = " ".join(last_tokens[:-1])
    last = re.sub(r"\W+", "", last).upper()
    first = re.sub(r"\W+", "", first).upper()
    if not last:
        return None
    return (last, first)


# ---------------------------------------------------------------------------
# Date extraction
# ---------------------------------------------------------------------------

_MD_RE = re.compile(r"(\d{1,2})/(\d{1,2})")


def extract_md(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    s = str(value).strip()
    if not s:
        return None
    if "DO NOT ASSIGN" in s.upper():
        return "NA"
    if "NEVER FLOAT" in s.upper():
        return None
    m = _MD_RE.search(s)
    if m:
        return f"{int(m.group(1))}/{int(m.group(2))}"
    return None


# ---------------------------------------------------------------------------
# Tracker loading
# ---------------------------------------------------------------------------

def build_simple_lookup(ws, name_col, start_date_col, start_row, end_row):
    """
    Build {normalized_name: 'M/D'} from a block of rows. Scans rightward
    from start_date_col to end-of-sheet and uses the rightmost non-empty
    date as the freshest entry (the tracker grows horizontally each cycle).
    Applies NAME_ALIASES on the tracker side so HC/Float-sheet spelling
    disagreements collapse to one canonical key.
    """
    table = {}
    last_col = ws.max_column
    for r in range(start_row, end_row + 1):
        name = ws.cell(row=r, column=name_col).value
        norm = normalize_name(name)
        if not norm:
            continue
        if norm in NAME_ALIASES:
            norm = NAME_ALIASES[norm]
        latest = None
        for c in range(start_date_col, last_col + 1):
            md = extract_md(ws.cell(row=r, column=c).value)
            if md is not None:
                latest = md
        if latest is not None and norm not in table:
            table[norm] = latest
    return table


def load_tracker(path, shift):
    wb = load_workbook(path, data_only=True)

    if shift == "noc":
        hc_sheet = "NOC RN HC LIST"
        float_sheet = "NOC RN FLOAT ROTATION"
    else:
        hc_sheet = "DAY RN HC LIST"
        float_sheet = "DAY RN FLOAT ROTATION"

    lookups = {"rn_hc": {}, "rn_float": {}, "cna_sit": {}, "cna_float": {}}

    if hc_sheet in wb.sheetnames:
        ws = wb[hc_sheet]
        lookups["rn_hc"] = build_simple_lookup(ws, 1, 3, 3, ws.max_row)

    if float_sheet in wb.sheetnames:
        ws = wb[float_sheet]
        lookups["rn_float"] = build_simple_lookup(ws, 1, 4, 4, ws.max_row)

    if "CNA ROTATION" in wb.sheetnames:
        ws = wb["CNA ROTATION"]
        sit_hdr = sitbreak_hdr = None
        for r in range(1, min(ws.max_row, 400) + 1):
            v = ws.cell(row=r, column=1).value
            if not v or not isinstance(v, str):
                continue
            up = v.strip().upper()
            if up == "CNA ROTATION LIST" or "IF CNA FLOATS" in up:
                continue
            if "SITTER BREAK ROTATION" in up:
                sitbreak_hdr = r
            elif "SITTER ROTATION" in up:
                sit_hdr = r

        def find_day_employee_header(after_row, before_row):
            for r in range(after_row, before_row):
                v = ws.cell(row=r, column=1).value
                if v and isinstance(v, str) and "EMPLOYEE-DAY" in v.upper():
                    return r
            return None

        def find_noc_employee_header(after_row, before_row):
            for r in range(after_row, before_row):
                v = ws.cell(row=r, column=1).value
                if v and isinstance(v, str) and "EMPLOYEE-NOC" in v.upper():
                    return r
            return None

        # CNA FLOAT — single shift-agnostic pool. Despite the header saying
        # "EMPLOYEE-DAY CNAs", this section contains ALL CNAs who float,
        # regardless of shift. Read the whole section for both DAY and NOC.
        float_day_hdr = find_day_employee_header(1, sit_hdr or ws.max_row)
        if float_day_hdr:
            float_end = (sit_hdr or ws.max_row) - 1
            lookups["cna_float"] = build_simple_lookup(
                ws, 1, 3, float_day_hdr + 1, float_end)

        # CNA SIT — has separate DAY and NOC sub-sections, use the right one.
        if sit_hdr:
            sit_day_hdr = find_day_employee_header(sit_hdr, sitbreak_hdr or ws.max_row)
            sit_end = (sitbreak_hdr or ws.max_row) - 1
            if sit_day_hdr:
                if shift == "noc":
                    noc_hdr = find_noc_employee_header(sit_day_hdr, sitbreak_hdr or ws.max_row)
                    if noc_hdr:
                        lookups["cna_sit"] = build_simple_lookup(
                            ws, 1, 3, noc_hdr + 1, sit_end)
                else:
                    noc_hdr = find_noc_employee_header(sit_day_hdr, sitbreak_hdr or ws.max_row)
                    day_end = (noc_hdr - 1) if noc_hdr else sit_end
                    lookups["cna_sit"] = build_simple_lookup(
                        ws, 1, 3, sit_day_hdr + 1, day_end)

    return lookups


# ---------------------------------------------------------------------------
# Roster reading
# ---------------------------------------------------------------------------

def get_value(v):
    return v.strip() if isinstance(v, str) else v


def detect_shift(ws, override=None):
    if override:
        return override
    for r in range(1, 14):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and "Coverage Period" in v:
            m = re.search(r"(\d{3,4})\s*-\s*(\d{3,4})", v)
            if m:
                start = int(m.group(1))
                return "noc" if start >= 1700 or start < 600 else "day"
    return "day"


def find_data_bounds(ws):
    header_row = None
    for r in range(1, 30):
        v = ws.cell(row=r, column=3).value
        if v and isinstance(v, str) and "Employee Name" in v:
            header_row = r
            break
    if header_row is None:
        header_row = 14
    data_start = header_row + 1
    data_end = data_start
    for r in range(data_start, ws.max_row + 1):
        name = ws.cell(row=r, column=3).value
        marker = ws.cell(row=r, column=1).value
        if (not name or not str(name).strip()) and \
           (not marker or not str(marker).strip() or str(marker).strip() != "*"):
            break
        data_end = r
    return header_row, data_start, data_end


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def unmerge_at(ws, row, col_letters):
    from openpyxl.utils import column_index_from_string
    targets = {column_index_from_string(c) for c in col_letters}
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row <= mr.max_row and mr.min_col in targets:
            to_remove.append(str(mr))
    for ref in to_remove:
        ws.unmerge_cells(ref)


def split_mo(ws, row):
    """Un-merge M:N:O at `row`, replace leftover MergedCells with Cells,
    copy M's style to O so dates render identically."""
    from copy import copy as _copy
    m_cell = ws.cell(row=row, column=13)
    style = {
        "font": _copy(m_cell.font),
        "alignment": _copy(m_cell.alignment),
        "border": _copy(m_cell.border),
        "fill": _copy(m_cell.fill),
        "number_format": m_cell.number_format,
    }
    unmerge_at(ws, row, ["M", "N", "O"])
    for col in (13, 14, 15):
        coord = (row, col)
        existing = ws._cells.get(coord)
        if isinstance(existing, MergedCell):
            ws._cells[coord] = Cell(ws, row=row, column=col)
    for col in (13, 15):
        c = ws.cell(row=row, column=col)
        c.font = style["font"]
        c.alignment = style["alignment"]
        c.border = style["border"]
        c.fill = style["fill"]
        c.number_format = style["number_format"]


def safe_set(ws, coord, value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        return
    cell.value = value


# ---------------------------------------------------------------------------
# Helper sheets
# ---------------------------------------------------------------------------

def add_rn_list_sheets(wb, decisions):
    """Append 'RN HC List' (#/HC/Name) and 'RN Float List' (Float/Name)."""
    rn_rows = [
        {"name": d["name"], "code": d["code"] or "",
         "hc": d["d1"] or "", "float": d["d2"] or ""}
        for d in decisions if d.get("profile") in RN_CLASS
    ]

    def date_key(val):
        if not val:
            return (1, datetime.max)
        try:
            m, day = str(val).strip().split("/")
            return (0, datetime(2026, int(m), int(day)))
        except Exception:
            return (1, datetime.max)

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=12, bold=True)

    hc_sorted = sorted(rn_rows, key=lambda r: date_key(r["hc"]))
    if "RN HC List" in wb.sheetnames:
        del wb["RN HC List"]
    s1 = wb.create_sheet("RN HC List")
    s1["A1"] = "RN-class — sorted by HC (ascending)"
    s1["A1"].font = title_font
    s1.merge_cells("A1:C1")
    s1["A1"].alignment = Alignment(horizontal="center")
    for col, h in enumerate(["#", "HC", "Name"], 1):
        c = s1.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    for i, r in enumerate(hc_sorted, 1):
        for col, v in enumerate([i, r["hc"], r["name"]], 1):
            c = s1.cell(row=3 + i, column=col, value=v)
            c.font = body_font
            c.border = border
            c.alignment = Alignment(horizontal="center" if col in (1, 2) else "left")
    for col, w in zip("ABC", [5, 8, 30]):
        s1.column_dimensions[col].width = w

    float_sorted = sorted(rn_rows, key=lambda r: date_key(r["float"]))
    if "RN Float List" in wb.sheetnames:
        del wb["RN Float List"]
    s2 = wb.create_sheet("RN Float List")
    s2["A1"] = "RN-class — sorted by Float (ascending)"
    s2["A1"].font = title_font
    s2.merge_cells("A1:B1")
    s2["A1"].alignment = Alignment(horizontal="center")
    for col, h in enumerate(["Float", "Name"], 1):
        c = s2.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    for i, r in enumerate(float_sorted, start=4):
        a = s2.cell(row=i, column=1, value=r["float"])
        b = s2.cell(row=i, column=2, value=r["name"])
        for c in (a, b):
            c.font = body_font
            c.border = border
        a.alignment = Alignment(horizontal="center")
        b.alignment = Alignment(horizontal="left")
    s2.column_dimensions["A"].width = 10
    s2.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build(roster_path, tracker_path, output_path, shift_override=None, verbose=False):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(roster_path, output_path)
    wb = load_workbook(output_path)

    keep = "Sheet2" if "Sheet2" in wb.sheetnames else wb.sheetnames[0]
    for sn in list(wb.sheetnames):
        if sn != keep:
            del wb[sn]
    ws = wb[keep]

    shift = detect_shift(ws, shift_override)
    lookups = load_tracker(tracker_path, shift)
    header_row, data_start, data_end = find_data_bounds(ws)
    if verbose:
        print(f"[info] header row {header_row}, data rows {data_start}-{data_end}, shift {shift.upper()}")

    employees = []
    for r in range(data_start, data_end + 1):
        name = get_value(ws.cell(row=r, column=3).value)
        prof = get_value(ws.cell(row=r, column=6).value)
        code = get_value(ws.cell(row=r, column=10).value)
        if not name:
            continue
        employees.append({"row": r, "name": name, "profile": prof, "code": code})

    def is_rn_class(p):
        return p in RN_CLASS or p in RN_NUMBER_NO_DATES
    def is_cna_class(p):
        return p in CNA_CLASS or p in CNA_NUMBER_NO_DATES

    rn_emps = [e for e in employees if is_rn_class(e["profile"])]
    cna_emps = [e for e in employees if is_cna_class(e["profile"])]
    other_emps = [e for e in employees if not is_rn_class(e["profile"]) and not is_cna_class(e["profile"])]

    decisions = []
    unmatched = []

    def process(emp, stream_counter, klass):
        prof = emp["profile"]
        code = (emp["code"] or "").strip().upper()
        norm = normalize_name(emp["name"])
        if norm in NAME_ALIASES:
            norm = NAME_ALIASES[norm]
        overrides = EMPLOYEE_OVERRIDES.get(norm, set()) if norm else set()

        number = None
        d1 = d2 = None

        if code in NO_NUMBER_CODES:
            pass
        elif prof in NO_TRACK:
            pass
        else:
            stream_counter[0] += 1
            number = stream_counter[0]
            if code in NO_DATES_CODES:
                pass
            elif prof in RN_NUMBER_NO_DATES or prof in CNA_NUMBER_NO_DATES:
                pass
            elif klass == "rn" and norm:
                attempted = []
                if "skip_left" not in overrides:
                    d1 = lookups["rn_hc"].get(norm)
                    attempted.append(d1)
                if "skip_right" not in overrides:
                    d2 = lookups["rn_float"].get(norm)
                    attempted.append(d2)
                if attempted and all(x is None for x in attempted):
                    unmatched.append((emp["name"], prof, "RN tables"))
            elif klass == "cna" and norm:
                attempted = []
                if "skip_left" not in overrides:
                    d1 = lookups["cna_sit"].get(norm)
                    attempted.append(d1)
                if "skip_right" not in overrides:
                    d2 = lookups["cna_float"].get(norm)
                    attempted.append(d2)
                if attempted and all(x is None for x in attempted):
                    unmatched.append((emp["name"], prof, "CNA tables"))

        decisions.append({"row": emp["row"], "number": number,
                          "d1": d1, "d2": d2, "name": emp["name"],
                          "code": emp["code"], "profile": prof})

    rn_counter = [0]
    for e in rn_emps:
        process(e, rn_counter, "rn")
    cna_counter = [0]
    for e in cna_emps:
        process(e, cna_counter, "cna")
    for e in other_emps:
        process(e, [0], "other")

    split_mo(ws, header_row)
    safe_set(ws, f"M{header_row}", "hc")
    safe_set(ws, f"O{header_row}", "float")

    for d in decisions:
        safe_set(ws, f"B{d['row']}", d["number"] if d["number"] is not None else None)
        split_mo(ws, d["row"])
        safe_set(ws, f"M{d['row']}", d["d1"] if d["d1"] else None)
        safe_set(ws, f"O{d['row']}", d["d2"] if d["d2"] else None)
        safe_set(ws, f"L{d['row']}", None)

    for r in range(data_end + 1, ws.max_row + 1):
        for col in range(1, 16):
            cell = ws.cell(row=r, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    add_rn_list_sheets(wb, decisions)
    wb.save(output_path)

    if verbose:
        print(f"[info] wrote {output_path}")
        if unmatched:
            print(f"[warn] {len(unmatched)} employees not found in tracker:")
            for name, prof, where in unmatched:
                print(f"        - {name!r} ({prof}) — searched {where}")

    return unmatched, shift


def main(argv=None):
    p = argparse.ArgumentParser(description="Build enriched daily nursing roster.")
    p.add_argument("--roster", required=True)
    p.add_argument("--tracker", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--shift", choices=["day", "noc"], default=None)
    p.add_argument("--verbose", action="store_true")
    args = p.parse_args(argv)
    build(args.roster, args.tracker, args.output, args.shift, args.verbose)


if __name__ == "__main__":
    main()
